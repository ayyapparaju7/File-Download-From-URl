import openpyxl
import wget
from tqdm import tqdm
from pathlib import Path

# Paths
excel_file_path = "C:\\File\\File.xlsx"
download_directory = "C:\\File\\Downloads"

# Ensure the download directory exists
Path(download_directory).mkdir(parents=True, exist_ok=True)

# Load the Excel file
try:
    wb_obj = openpyxl.load_workbook(excel_file_path)
    sheet_obj = wb_obj.active
except Exception as e:
    print(f"Error loading Excel file: {e}")
    exit()

# Iterate over the rows in the Excel sheet
total_rows = sheet_obj.max_row
success_count = 0
failure_count = 0
failed_downloads = []

for i in tqdm(range(1, total_rows + 1), desc="Downloading files"):
    filename = sheet_obj.cell(row=i, column=1).value
    downloadUrl = sheet_obj.cell(row=i, column=2).value

    # Ensure filename and downloadUrl are not None
    if not filename or not downloadUrl:
        failed_downloads.append((filename, downloadUrl, "Missing filename or URL"))
        failure_count += 1
        continue

    # Ensure the file is saved in the specified download directory
    file_path = Path(download_directory) / filename

    try:
        wget.download(downloadUrl, str(file_path))
        success_count += 1
    except Exception as e:
        failed_downloads.append((filename, downloadUrl, str(e)))
        failure_count += 1


if failed_downloads:
    print("Failed downloads:")
    for failure in failed_downloads:
        print(f"  Filename: {failure[0]} - URL: {failure[1]} - Reason: {failure[2]}")

# Display results
print("\nDownload Complete")
print(f"Total files processed: {total_rows}")
print(f"Total successful downloads: {success_count}")
print(f"Total failed downloads: {failure_count}\n")
